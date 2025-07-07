import os
import re
import zipfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from tabulate import tabulate

# Input Excel directory
excel_dir = "excel_files"

# Create Excel folder if missing
if not os.path.exists(excel_dir):
    os.makedirs(excel_dir)
    print(f"Directory '{excel_dir}' created. Please add Excel files and rerun the script.")
    exit()

# Version parser from filename
def extract_version(filename):
    match = re.search(r'V(\d+)\.(\d+)\.(\d+)', filename)
    if match:
        return tuple(map(int, match.groups()))
    return (0, 0, 0)

# List Excel files
excel_files = [f for f in os.listdir(excel_dir) if f.endswith('.xlsx')]
excel_files.sort(key=extract_version)

if len(excel_files) < 2:
    print("Not enough versioned Excel files found in the directory.")
    exit()

# Pick latest and previous
file_old = os.path.join(excel_dir, excel_files[-2])
file_new = os.path.join(excel_dir, excel_files[-1])

# Create output folder
output_dir = "Output"
os.makedirs(output_dir, exist_ok=True)

# Save to file with timestamp
timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
output_excel = os.path.join(output_dir, f"difference_{timestamp}.xlsx")
summary_txt = os.path.join(output_dir, f"summary_{timestamp}.txt")
artifact_zip = os.path.join(output_dir, f"artifact_{timestamp}.zip")

print(f"Old File: '{file_old}'")
print(f"New File: '{file_new}'")

# Load Excel files
wb_old = load_workbook(file_old)
wb_new = load_workbook(file_new)

# Highlight style
highlight = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
bold_font = Font(bold=True, color='FFFFFF')

# Track changes - detailed view
summary_data = []
detailed_changes = []

print("Processing sheets...")

for sheet_name in wb_old.sheetnames:
    if sheet_name not in wb_new.sheetnames:
        summary_data.append((sheet_name, 'Missing in new file'))
        detailed_changes.append([sheet_name, "-", "-", "-", "Sheet missing in new file"])
        continue

    ws_old = wb_old[sheet_name]
    ws_new = wb_new[sheet_name]
    max_row = max(ws_old.max_row, ws_new.max_row)
    max_col = max(ws_old.max_column, ws_new.max_column)

    changes = 0
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val_old = ws_old.cell(row=row, column=col).value
            val_new = ws_new.cell(row=row, column=col).value
            if val_old != val_new:
                # Highlight changed cell in new file
                ws_new.cell(row=row, column=col).fill = highlight
                ws_new.cell(row=row, column=col).font = bold_font
                changes += 1
                detailed_changes.append([
                    sheet_name,
                    ws_new.cell(row=row, column=col).coordinate,
                    val_old,
                    val_new,
                    "Changed" if val_old is not None and val_new is not None else ("Added" if val_old is None else "Removed")
                ])
    summary_data.append((sheet_name, f"{changes} changes" if changes else "No changes"))

# Also detect new sheets in the new file
for sheet_name in wb_new.sheetnames:
    if sheet_name not in wb_old.sheetnames and sheet_name != "Summary_of_Changes":
        summary_data.append((sheet_name, 'New sheet in new file'))
        detailed_changes.append([sheet_name, "-", "-", "-", "Sheet added in new file"])

# Summary sheet
ws_summary = wb_new.create_sheet("Summary_of_Changes")
ws_summary.append(["Sheet Name", "Change Summary"])
for entry in summary_data:
    ws_summary.append(entry)

# Save Excel output
wb_new.save(output_excel)

# Save summary to text file (detailed, cell-level)
with open(summary_txt, 'w', encoding='utf-8') as f:
    f.write(f"Comparison Summary - {timestamp}\n")
    f.write(f"Old File: {file_old}\n")
    f.write(f"New File: {file_new}\n\n")
    f.write(tabulate(summary_data, headers=["Sheet Name", "Change Summary"], tablefmt="grid"))
    f.write('\n\nDETAILED DIFFERENCES:\n')
    if detailed_changes:
        f.write(tabulate(
            detailed_changes,
            headers=["Sheet", "Cell", "Old Value", "New Value", "Change Type"],
            tablefmt="grid"
        ))
    else:
        f.write("No cell-level differences found.\n")

# Create artifact zip (only latest files, never zip any ZIP file)
with zipfile.ZipFile(artifact_zip, 'w') as zipf:
    zipf.write(output_excel, os.path.basename(output_excel))
    zipf.write(summary_txt, os.path.basename(summary_txt))

print("Summary of Changes:")
print(tabulate(summary_data, headers=["Sheet Name", "Change Summary"], tablefmt="grid"))
print(f"Excel output saved to: {output_excel}")
print(f"Summary saved to: {summary_txt}")
print(f"Artifact created at: {artifact_zip}")
