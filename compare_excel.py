from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from tabulate import tabulate

file_old = "excel_files/Telstra Acquisitons - Decision Harvesting Workbook V1.7.0.xlsx"
file_new = "excel_files/Telstra Acquisitons - Decision Harvesting Workbook V1.8.0.xlsx"
output_file = 'highlighted_all_sheets.xlsx'

print(f" Old File is '{file_old}'.")
print(f" New File is '{file_new}'.")

# Load both workbooks
wb_old = load_workbook(file_old)
wb_new = load_workbook(file_new)

print(" Loaded workbooks.")

# Create a fill style for highlighting
highlight = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='lightDown')
bold_font = Font(bold=True,color='FFFFFF')

# Track changes for summary
summary_data = []

print(" Processing Sheets in workbooks.")

for sheet_name in wb_old.sheetnames:
    if sheet_name not in wb_new.sheetnames:
        summary_data.append((sheet_name, 'Missing in new file'))
        continue

    ws_old = wb_old[sheet_name]
    ws_new = wb_new[sheet_name]

    max_row = max(ws_old.max_row, ws_new.max_row)
    max_col = max(ws_old.max_column, ws_new.max_column)

    changes = 0
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val_old = ws_old.cell(row=row, column=col).value
            val_new = ws_new.cell(row=row, column=col).value
            if val_old != val_new:
                ws_new.cell(row=row, column=col).fill = highlight
                ws_new.cell(row=row, column=col).font = bold_font
                changes += 1

    if changes:
        summary_data.append((sheet_name, f"{changes} changes"))
    else:
        summary_data.append((sheet_name, "No changes"))

# Add summary sheet
ws_summary = wb_new.create_sheet("Summary_of_Changes")
ws_summary.append(["Sheet Name", "Change Summary"])
for entry in summary_data:
    ws_summary.append(entry)

# Print summary table in console
print(" Summary of Changes: ****************")
print(tabulate(summary_data, headers=["Sheet Name", "Change Summary"], tablefmt="grid"))

# Save output
wb_new.save(output_file)
print(f" Done! Differences highlighted and summary added in '{output_file}'.")
